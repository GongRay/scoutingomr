# USAGE
# python test_grader.py --image images/test_01.png

# import the necessary packages
from imutils.perspective import four_point_transform
from imutils import contours
import numpy as np
import argparse
import imutils
import cv2
import glob
import shutil
import openpyxl
import os.path
import qrtools
from pyzbar.pyzbar import decode
from os import path
from openpyxl import Workbook

def display(im, bbox):
    n = len(bbox)
    for j in range(n):
        cv2.line(im, tuple(bbox[j][0]), tuple(bbox[ (j+1) % n][0]), (255,0,0), 3)

if path.exists("Results.xlsx"):
        wb = openpyxl.load_workbook('Results.xlsx')
else:
        wb = Workbook()

for image_file in glob.iglob('images/*.png'):
        matchNum = 0
        stationPos = ""
        qrDecoder = cv2.QRCodeDetector()
        # load the image, convert it to grayscale, blur it
        # slightly, then find edges
        image = cv2.imread(image_file)
        # Detect and decode the qrcode
        data,bbox,rectifiedImage = qrDecoder.detectAndDecode(image)
    
        if len(data)>0:
            split = data.split(':')
            print(split)
            matchNum = int(split[0])
            display(image, bbox)
            rectifiedImage = np.uint8(rectifiedImage);              
        else:
            split = 'NA'
            print("QR Code not detected")

        if split[1] is '1':
            stationPos = 'Red 1'
        elif split[1] is '2':
            stationPos = 'Red 2'
        elif split[1] is '3':
            stationPos = 'Red 3'
        elif split[1] is '4':
            stationPos = 'Blue 1'
        elif split[1] is '5':
            stationPos = 'Blue 2'
        elif split[1] is '6':
            stationPos = 'Blue 3'
        else:
            stationPos = split[1]
            
        cv2.rectangle(image,(565,30),(685,150),(255,255,255),-1)
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        blurred = cv2.GaussianBlur(gray, (5, 5), 0)

        # find contours in the edge map, then initialize
        # the contour that corresponds to the document
        cnts = cv2.findContours(blurred.copy(), cv2.RETR_EXTERNAL,
                cv2.CHAIN_APPROX_SIMPLE)
        cnts = imutils.grab_contours(cnts)
        docCnt = None

        # ensure that at least one contour was found
        if len(cnts) > 0:
                # sort the contours according to their size in
                # descending order
                cnts = sorted(cnts, key=cv2.contourArea, reverse=True)

                # loop over the sorted contours
                for c in cnts:
                        # approximate the contour
                        peri = cv2.arcLength(c, True)
                        approx = cv2.approxPolyDP(c, 0.02 * peri, True)

                        # if our approximated contour has four points,
                        # then we can assume we have found the paper
                        if len(approx) == 4:
                                docCnt = approx
                                break

        # apply a four point perspective transform to both the
        # original image and grayscale image to obtain a top-down
        # birds eye view of the paper
        paper = four_point_transform(image, docCnt.reshape(4, 2))
        warped = four_point_transform(gray, docCnt.reshape(4, 2))

        # apply Otsu's thresholding method to binarize the warped
        # piece of paper
        thresh = cv2.threshold(warped, 0, 255,
                cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU)[1]

        # find contours in the thresholded image, then initialize
        # the list of contours that correspond to questions
        cnts = cv2.findContours(thresh.copy(), cv2.RETR_EXTERNAL,
                cv2.CHAIN_APPROX_SIMPLE)
        cnts = imutils.grab_contours(cnts)
        questionCnts = []
        cv2.imshow("Threshold", thresh)
        # loop over the contours
        for c in cnts:
                # compute the bounding box of the contour, then use the
                # bounding box to derive the aspect ratio
                (x, y, w, h) = cv2.boundingRect(c)
                ar = w / float(h)

                # in order to label the contour as a question, region
                # should be sufficiently wide, sufficiently tall, and
                # have an aspect ratio approximately equal to 1
                if w >= 20 and h >= 20 and ar >= 0.95 and ar <= 1.05:
                        questionCnts.append(c)

        # sort the question contours top-to-bottom, then initialize
        questionCnts = contours.sort_contours(questionCnts,
                method="top-to-bottom")[0]

        # each question has 5 possible answers, to loop over the
        # question in batches of 5
        teamnum = ""
        for (q, i) in enumerate(np.arange(0, 40, 10)):
                cnts = contours.sort_contours(questionCnts[i:i + 10])[0]
                bubbled = None

                for (j, c) in enumerate(cnts):
                        mask = np.zeros(thresh.shape, dtype="uint8")
                        cv2.drawContours(mask, [c], -1, 255, -1)
                        
                        mask = cv2.bitwise_and(thresh, thresh, mask=mask)
                        total = cv2.countNonZero(mask)

                        if bubbled is None or total > bubbled[0]:
                                bubbled = (total, j)
                                
                teamnum = teamnum + str(bubbled[1])
            
        if not str(int(teamnum)) in wb.sheetnames:
                ws = wb.create_sheet(str(int(teamnum)))
                ws.append(['Team #', int(teamnum)])
                ws.append(['Match #', 'Driver Station', 'Start Pos', 'Auto Move', 'Auto High Hot', 'Auto High Cold', 'Auto Missed', 'Inbound from Human', 'Picked Up', 'Truss To Human', 'Truss', 'Caught', 'Missed Shot', 'Missed 2 or more times', 'Scored High', 'Scored Low']) 
        else:
                ws = wb.get_sheet_by_name(str(int(teamnum)))
                
        for (q, i) in enumerate(np.arange(40, 45, 5)):
                cnts = contours.sort_contours(questionCnts[i:i + 5])[0]
                bubbled = None

                for (j, c) in enumerate(cnts):
                        mask = np.zeros(thresh.shape, dtype="uint8")
                        cv2.drawContours(mask, [c], -1, 255, -1)

                        mask = cv2.bitwise_and(thresh, thresh, mask=mask)
                        total = cv2.countNonZero(mask)

                        if bubbled is None or total > bubbled[0]:
                                bubbled = (total, j)
                                
                if bubbled[1] is 0:
                    startPos = 'Left'
                elif bubbled[1] is 1:
                    startPos = 'Center'
                elif bubbled[1] is 2:
                    startPos = 'Right'
                elif bubbled[1] is 3:
                    startPos = 'GZ'
                elif bubbled[1] is 4:
                    startPos = 'No Show'
                else:
                    startPos = bubbled[1]

        for (q, i) in enumerate(np.arange(45, 46, 1)):
                cnts = contours.sort_contours(questionCnts[i:i + 1])[0]
                bubbled = None

                for (j, c) in enumerate(cnts):
                        mask = np.zeros(thresh.shape, dtype="uint8")
                        cv2.drawContours(mask, [c], -1, 255, -1)
                        
                        mask = cv2.bitwise_and(thresh, thresh, mask=mask)
                        total = cv2.countNonZero(mask)

                        if bubbled is None or total > bubbled[0]:
                                bubbled = (total, 'Yes')
                                
                        if bubbled[0] < 200 and total < 200:
                                bubbled = (total, 'NA')
                                
                autoMove = bubbled[1]                

        for (q, i) in enumerate(np.arange(46, 55, 3)):
                cnts = contours.sort_contours(questionCnts[i:i + 3])[0]
                bubbled = None
                for (j, c) in enumerate(cnts):
                        
                        mask = np.zeros(thresh.shape, dtype="uint8")
                        cv2.drawContours(mask, [c], -1, 255, -1)

                        mask = cv2.bitwise_and(thresh, thresh, mask=mask)
                        total = cv2.countNonZero(mask)
                                
                        if bubbled is None or total > bubbled[0]:
                                bubbled = (total, j+1)

                        if j == 2 and bubbled[0] < 200 and total < 200:
                                bubbled = (total, 0)

                if q is 0:
                        autoHighHot = bubbled[1]
                elif q is 1:
                        autoHighCold = bubbled[1]
                elif q is 2:
                        autoMiss = bubbled[1]
        humanInbound = 0                
        pickedUp = 0
        trussToHuman = 0
        truss = 0
        caught = 0
        missed = 0
        missed2 = 0
        scoredHigh = 0
        scoredLow = 0
        for (q, i) in enumerate(np.arange(55, 118, 9)):
                cnts = contours.sort_contours(questionCnts[i:i + 9])[0]
                bubbled = None

                for (j, c) in enumerate(cnts):
                        mask = np.zeros(thresh.shape, dtype="uint8")
                        cv2.drawContours(mask, [c], -1, 255, -1)

                        mask = cv2.bitwise_and(thresh, thresh, mask=mask)
                        total = cv2.countNonZero(mask)
                        
                       # if bubbled is None or total > bubbled[0]-20:
                       #         bubbled = (total, j)
                        if j is 0 and total >= 250:
                            humanInbound += 1
                        elif j is 1 and total >= 250:
                            pickedUp += 1
                        elif j is 2 and total >= 250:
                            trussToHuman += 1
                        elif j is 3 and total >= 250:
                            truss += 1
                        elif j is 4 and total >= 250:
                            caught += 1
                        elif j is 5 and total >= 250:
                            missed += 1
                        elif j is 6 and total >= 250:
                            missed2 += 1
                        elif j is 7 and total >= 250:
                            scoredHigh += 1
                        elif j is 8 and total >= 250:
                            scoredLow += 1
                        #if j == 8 and bubbled[0] < 200 and total < 200:
                        #        bubbled = (total, 'NA')
                                
        ws.append([matchNum, stationPos, startPos, autoMove, autoHighHot, autoHighCold, autoMiss, humanInbound, pickedUp, trussToHuman, truss, caught, missed, missed2, scoredHigh, scoredLow])
        wb.save('Results.xlsx')
        #if not path.exists('images/Processed/'+str(matchNum)+'/') and matchNum is not 0:
        #    os.makedirs('images/Processed/'+str(matchNum)+'/')
        #    shutil.move(image_file, 'images/Processed/'+str(matchNum)+'/')
cv2.waitKey(0)
